import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color
from openpyxl.styles.colors import BLUE
from datetime import datetime
import re
import json

# Function to read the GitHub access token from a file
def read_token_from_file(file_path):
    with open(file_path, 'r') as file:
        return file.read().strip()
token_file_path = "C:\\Users\\wquraishi\\Documents\\GitHub-Config\/github_token.txt"
access_token = read_token_from_file(token_file_path)


def sanitize_for_excel(text):
    if not isinstance(text, str):
        text = str(text)
    if not text:
        return text  # or return "" to avoid NoneType issues
    # Remove HTML tags
    text = re.sub('<[^<]+?>', ' ', text)
    # Replace URLs with a simple placeholder or remove them
    text = re.sub(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', '[LINK]', text)
    # Remove or replace other illegal characters as needed
    text = text.replace('\n', ' ').replace('\r', '').replace('\t', ' ')
    # Truncate to avoid Excel cell character limit issues
    text = (text[:32767]) if len(text) > 32767 else text
    return text

def fetch_all_items(base_url, headers):
    items = []
    page = 1
    while True:
        # Construct the full URL with query parameters for each request
        api_url = f"{base_url}?state=all&page={page}&per_page=100"
        response = requests.get(api_url, headers=headers)
        print(f"Fetching {api_url}")  # Debug print to check the constructed URL
        
        if response.status_code == 200:
            data = response.json()
            if not data:
                break  # No more data, exit the loop
            items.extend(data)
            page += 1
        else:
            print(f"Failed to fetch data. Status Code: {response.status_code}. Response: {response.text} status_code: {response.status_code}")
            break
    return items

def get_username_from_string(input_string):
   start_index = input_string.find("'login': '") + len("'login': '")
   end_index = input_string.find("_kpmg'")
   input_string = input_string[start_index:end_index]
   return input_string

#kpmg-global-technology-and-knowledge/digital-matrix-app
repo_owner = "kpmg-global-technology-and-knowledge"
repo_name = "digital-matrix-app"
#repo_owner = "microsoft"
#repo_name = "azurechat"

headers = {"Authorization": f"Bearer {access_token}"}

issues_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/issues"
pulls_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/pulls"

# Fetch issues and pull requests with pagination
#issues_data = fetch_all_items(issues_url, headers)
pulls_data = fetch_all_items(pulls_url, headers)

# The rest of your script remains the same...

# Create an Excel workbook
wb = Workbook()
sheet = wb.active
sheet.title = "Pull R. for digital-matrix-app"
# Set header row font to bold ...
# Continue as before.

# Write the header row
header_row = ["Number", "Type", "Title", "Body", "Reporter (User)", "Labels","Milestone","State", "Reviewers", "Committers"]
for col_num, header in enumerate(header_row, 1):
    col_letter = get_column_letter(col_num)
    sheet[f"{col_letter}1"] = header

# Write pull requests data to the Excel file
for pull_num, pull in enumerate(pulls_data, 2):
    pull_number = pull["number"]
    pull_type = "Pull Request"
    pull_title = sanitize_for_excel(pull["title"])
    pull_body = pull["body"]
    pull_assignees = ",".join(assignee["login"] for assignee in pull["assignees"])
    pull_labels = ",".join(label["name"] for label in pull["labels"])
    pull_milestone = pull["milestone"]["title"] if pull["milestone"] else ""
    pull_state = pull["state"]
    pull_url = pull["html_url"]
    pull_reviewers = ",".join(reviewer["login"] for reviewer in pull["requested_reviewers"])
    pull_committers = pull["user"]["login"]
#try:
    cell = f"A{pull_num}"
    sheet[cell].hyperlink = f'{pull_url}'
    sheet[cell].value = pull_number
    sheet[cell].font = Font(color=Color(rgb=BLUE))
    sheet[f"B{pull_num}"] = pull_type
    sheet[f"C{pull_num}"] = pull_title
    sheet[f"D{pull_num}"] = pull_body
    sheet[f"E{pull_num}"] = pull_assignees
    sheet[f"F{pull_num}"] = pull_labels
    sheet[f"G{pull_num}"] = pull_milestone
    sheet[f"H{pull_num}"] = pull_state
    sheet[f"I{pull_num}"] = pull_reviewers
    sheet[f"J{pull_num}"] = pull_committers
#except Exception as e:
    # Catch any exception
#    print(f"An error occurred: {e}")    

# Adjust column widths
for col in sheet.columns:
    max_length = 0
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[col[0].column_letter].width = adjusted_width

# Save the workbook as an Excel file
if pulls_data:
    print("Data fetched, writing to Excel...")
else:
    print("No data fetched, please check the fetch logic.")

# Generate output filename with current datetime suffix
current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
output_filename = f"Pull_requests_{current_datetime}.xlsx"
wb.save(output_filename)
