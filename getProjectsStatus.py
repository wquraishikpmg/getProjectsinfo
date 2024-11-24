import json
from datetime import datetime
import pandas as pd
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# Function to read the token from a file
def read_token_from_file(file_path):
    with open(file_path, 'r') as file:
        return file.read().strip()

# Path to the token file
token_file_path = 'github_token.txt'

# Read the token from the file
token = read_token_from_file(token_file_path)

# Define project numbers and their corresponding titles
project_mapping = {
    12: "Workbench Program Status",
    18: "Workbench-Platform-Americas-Streams",
    20: "Workbench-Platform-ASPAC-Streams",
    27: "Workbench-Platform-EMEA-Streams"
}

# Shortened sheet names to fit Excel's 31-character limit
shortened_project_mapping = {number: title[:31] for number, title in project_mapping.items()}

# Initialize a dictionary to hold DataFrames for each project
project_dataframes = {}
processed_data = {}  # Dictionary to hold preprocessed data for each project

# Corrected query template with proper handling of after cursor
query_template = '''
{
  organization(login: "kpmg-global-technology-and-knowledge") {
    projectV2(number: %d) {
      items(first: 100, after: %s) {
        pageInfo {
          endCursor
          hasNextPage
        }
        nodes {
          content {
            ... on Issue {
              id
              number
              title
              url
              createdAt
              updatedAt
              state
              author {
                login
              }
              labels(first: 10) {
                nodes {
                  name
                }
              }
              milestone {
                title
              }
            }
          }
          fieldValues(first: 100) {
            nodes {
              ... on ProjectV2ItemFieldValueCommon {
                field {
                  ... on ProjectV2FieldCommon {
                    name
                  }
                }
              }
              ... on ProjectV2ItemFieldTextValue {
                field {
                  ... on ProjectV2Field {
                    name
                  }
                }
                text
              }
              ... on ProjectV2ItemFieldSingleSelectValue {
                field {
                  ... on ProjectV2SingleSelectField {
                    name
                  }
                }
                name
              }
            }
          }
        }
      }
    }
  }
}
'''

# Function to extract status from field values
def extract_status(field_values):
    for field in field_values:
        if 'field' in field:
            print(f"Processing field: {field}")  # DEBUG: Print each field being processed
            if field['field']['name'] == 'Status':
                if 'text' in field:
                    return field['text']
                elif 'name' in field:
                    return field['name']
    return None

# Function to fetch all issues for a project, handling pagination
def fetch_all_issues_for_project(project_number):
    issues = []
    end_cursor = None  # Initialize cursor as None
    has_next_page = True
  
    while has_next_page:
        cursor_str = f'"{end_cursor}"' if end_cursor else 'null'  # Proper handling of cursor value in query
        query = query_template % (project_number, cursor_str)
        response = requests.post("https://api.github.com/graphql", json={'query': query}, headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"})
        data = response.json()
        
        # Check for and handle errors in the response
        if "errors" in data:
            print(f"Error fetching data for project {project_number}: {data['errors']}")
            break

        # Parse the JSON response and extract issues
        page_info = data['data']['organization']['projectV2']['items']['pageInfo']
        nodes = data['data']['organization']['projectV2']['items']['nodes']
        
        for node in nodes:
            issue = node['content']
            if issue:
                issue_number = issue['number']
                title = issue['title']
                url = issue['url']
                created_at = issue['createdAt']
                updated_at = issue['updatedAt']
                state = issue['state']
                author = issue['author']['login']
                labels = ", ".join([label['name'] for label in issue['labels']['nodes']])
                milestone = issue['milestone']['title'] if issue.get('milestone') else None
                print(f"Fetching status for issue: {issue_number}")  # DEBUG: Verify issue number being processed
                status = extract_status(node['fieldValues']['nodes'])  # Extract status field value
                issues.append({
                    'Title': title,
                    'URL': url,
                    'Created At': created_at,
                    'Updated At': updated_at,
                    'State': state,
                    'Author': author,
                    'Labels': labels,
                    'Milestone': milestone,
                    'Status': status  # Ensure column name is 'Status'
                })

        has_next_page = page_info['hasNextPage']
        end_cursor = page_info['endCursor']

    print(f"Issues fetched for project {project_number}: {issues}")  # DEBUG: Print all issues fetched
    return issues

# Fetch all issues for each project and store in DataFrames
for project_number, project_title in project_mapping.items():
    issues = fetch_all_issues_for_project(project_number)
    df = pd.DataFrame(issues)
    df['Status'] = df['Status'].astype(str)  # Ensure the Status column type is string
    print(f"DataFrame for project {project_number}:\n{df.head()}")  # DEBUG: Check DataFrame content
    print(f"DataFrame columns: {df.columns}")  # DEBUG: Verify columns
    project_dataframes[shortened_project_mapping[project_number]] = df

    # Pre-process data into dictionaries for access without indexing issues
    processed_data[shortened_project_mapping[project_number]] = df.to_dict(orient='records')

# Create a timestamped Excel writer
current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
output_filename = f"getProjectsStatus_{current_datetime}.xlsx"

with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    for project_title, df in project_dataframes.items():
        sheet_name = project_title
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# Load the newly created Excel file to verify it
workbook = load_workbook(output_filename)

# Define the formulas for the additional columns
label_status_formula = '''=IFERROR(MID(G2, SEARCH("Status: ", G2) + LEN("Status: "), IF(ISNUMBER(SEARCH(",", G2, SEARCH("Status: ", G2) + LEN("Status: "))), SEARCH(",", G2, SEARCH("Status: ", G2) + LEN("Status: ")) - (SEARCH("Status: ", G2) + LEN("Status: ")), LEN(G2))), "")'''
issuetype_formula = '''=IF(OR(UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "FEATURE", UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "USER STORY", UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "TASK", UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "EPIC", UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "OPERATIONAL", UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "DEFECT"), UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)), IF(OR(LEFT(G2, 4) = "Pod:", G2 = ""), "", IFERROR(IF(ISERROR(FIND(",", G2)), G2, LEFT(G2, FIND(",", G2) - 1)), G2)))'''
pod_formula = '''=IFERROR(MID(G2, SEARCH("Pod: ", G2) + LEN("Pod: "), SEARCH(",", G2, SEARCH("Pod: ", G2)) - (SEARCH("Pod: ", G2) + LEN("Pod: "))), "")'''
isdefect_formula = '''=IF(ISNUMBER(SEARCH("Defect", G2)), "Defect", "")'''
convertHyperlink = '''=HYPERLINK(B2, TRIM(RIGHT(SUBSTITUTE(B2, "/", REPT(" ", 100)), 100)))'''
font = Font(color="0000FF", underline="single")

# Ensure each sheet in the workbook is processed
for project_title in workbook.sheetnames:
    if project_title in processed_data:
        sheet = workbook[project_title]
        sheet.sheet_state = 'visible'  # Ensure each sheet is visible

        # Add the headers for new columns
        sheet['H1'] = 'LabelStatus'
        sheet['I1'] = 'LabelIssueType'
        sheet['J1'] = 'Pod'
        sheet['K1'] = 'IsDefect'
        sheet['L1'] = 'Milestone'
        sheet['M1'] = 'GitHub Link '
        sheet['N1'] = 'POD Project '
        sheet['O1'] = 'Status'  # Header for Status column

        # Add the formulas to the respective columns for each row of data
        for row, record in enumerate(processed_data[project_title], start=2):
            sheet[f'H{row}'] = label_status_formula.replace('G2', f'G{row}')
            sheet[f'I{row}'] = issuetype_formula.replace('G2', f'G{row}')
            sheet[f'J{row}'] = pod_formula.replace('G2', f'G{row}')
            sheet[f'K{row}'] = isdefect_formula.replace('G2', f'G{row}')
            sheet[f'M{row}'] = convertHyperlink.replace('B2', f'B{row}')
            sheet[f'M{row}'].font = font
            sheet[f'N{row}'] = project_title

            sheet[f'L{row}'] = record.get('Milestone', '')
            status_value = record.get('Status', '')
            print(f"Writing Status '{status_value}' to row {row}")  # DEBUG: Print status value being written
            sheet[f'O{row}'] = status_value

# Ensure only one active sheet and it's visible
workbook.active = workbook.worksheets[0]
workbook.active.sheet_state = 'visible'

# Save the updated workbook
workbook.save(output_filename)

print(f"Issues successfully written to {output_filename} with additional columns including 'Status'.")
