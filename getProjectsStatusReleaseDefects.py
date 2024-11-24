import subprocess
import json
from datetime import datetime
import pandas as pd
import re
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font


# Function to read the token from a file
def read_token_from_file(file_path):
    with open(file_path, 'r') as file:
        return file.read().strip()

# Path to the token file
token_file_path = 'github_token.txt'

# Read the token from the file
token = read_token_from_file(token_file_path)

# Define project numbers and their corresponding titles
project_mapping = {
    12: "Workbench Program Status",
    18: "Workbench-Platform-Americas-Streams",
    20: "Workbench-Platform-ASPAC-Streams",
    27: "Workbench-Platform-EMEA-Streams"
}

# Shortened sheet names to fit Excel's 31-character limit
shortened_project_mapping = {number: title[:31] for number, title in project_mapping.items()}

# Initialize a dictionary to hold DataFrames for each project
project_dataframes = {}
processed_data = {}  # Dictionary to hold preprocessed data for each project

# Query template for fetching project issues
# Corrected query template with proper handling of after cursor
query_template = '''
{
  organization(login: "kpmg-global-technology-and-knowledge") {
    projectV2(number: %d) {
      items(first: 100, after: %s) {  # CHANGED: Correct handling of the cursor value
        pageInfo {
          endCursor
          hasNextPage
        }
        nodes {
          content {
            ... on Issue {
              id
              number
              title
              url
              createdAt
              updatedAt
              state
              author {
                login
              }
              labels(first: 10) {
                nodes {
                  name
                }
              }
              milestone {
                title
              }
            }
          }
          fieldValues(first: 100) {
            nodes {
              ... on ProjectV2ItemFieldValueCommon {
                field {
                  ... on ProjectV2FieldCommon {
                    name
                  }
                }
              }
              ... on ProjectV2ItemFieldTextValue {
                field {
                  ... on ProjectV2Field {
                    name
                  }
                }
                text
              }
              ... on ProjectV2ItemFieldSingleSelectValue {
                field {
                  ... on ProjectV2SingleSelectField {
                    name
                  }
                }
                name
              }
            }
          }
        }
      }
    }
  }
}
'''

# Function to extract status from field values
def extract_status(field_values):
    for field in field_values:
        if 'field' in field:
            #print(f"Processing field: {field}")  # DEBUG: Print each field being processed
            if field['field']['name'] == 'Status':
                if 'text' in field:
                    return field['text']
                elif 'name' in field:
                    return field['name']
    return None
# BEGIN of Function call to add Defects
def fetch_defects_content(file_path):
    # Read the Excel file
    df = pd.read_excel(file_path, sheet_name='Defects')
    
    # Convert the defects dataframe to a string
    defects_content = "## Defects\n\n"
    for index, row in df.iterrows():
        defect_description = " - ".join([str(item) for item in row if pd.notna(item)])
        defects_content += f"* {defect_description}\n"
    
    return defects_content
# End of Function call to add Defects    
# Function to fetch all issues for a project, handling pagination
def fetch_all_issues_for_project(project_number):
    issues = []
    end_cursor =  None  # CHANGED: Initialize cursor as None
    has_next_page = True
    
    while has_next_page:
        cursor_str = f'"{end_cursor}"' if end_cursor else 'null'  # CHANGED: Proper handling of cursor value in query
        query = query_template % (project_number, cursor_str)
        #curl_command = [
        #    "curl",
        #    "-H", f"Authorization: Bearer {token}",
        #    "-H", "Content-Type: application/json",
        #    "-X", "POST",
        #    "--data", json.dumps({"query": query}),
        #    "https://api.github.com/graphql"
        #]
        # Define the GitHub API URL
        url = "https://api.github.com/graphql"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        response = requests.post(url, json={'query': query}, headers=headers)
        data=response.json()
        #result = subprocess.run(curl_command, capture_output=True, text=True)
        #data = json.loads(result.stdout)
        
        # Check for and handle errors in the response
        if "errors" in data:
            print(f"Error fetching data for project {project_number}: {data['errors']}")
            break

        # Parse the JSON response and extract issues
        page_info = data['data']['organization']['projectV2']['items']['pageInfo']
        nodes = data['data']['organization']['projectV2']['items']['nodes']
        
        for node in nodes:
            issue = node['content']
            if issue:
                issue_number = issue['number']
                title = issue['title']
                url = issue['url']
                created_at = issue['createdAt']
                updated_at = issue['updatedAt']
                state = issue['state']
                author = issue['author']['login']
                labels = ", ".join([label['name'] for label in issue['labels']['nodes']])
                milestone = issue['milestone']['title'] if issue.get('milestone') else None
                #print(f"Fetching status for issue: {issue_number}")  # DEBUG: Verify issue number being processed
                status = extract_status(node['fieldValues']['nodes'])
                #print(f"Fetching status value : {status}")
                issues.append({
                    'Title': title,
                    'URL': url,
                    'Created At': created_at,
                    'Updated At': updated_at,
                    'State': state,
                    'Author': author,
                    'Labels': labels,
                    'Milestone': milestone,
                    'Status': status
                })

        has_next_page = page_info['hasNextPage']
        end_cursor = page_info['endCursor']
    #print(f"Issues fetched for project {project_number}: {issues}") 
    return issues

# Fetch all issues for each project and store in DataFrames
for project_number, project_title in project_mapping.items():
    issues = fetch_all_issues_for_project(project_number)
    df = pd.DataFrame(issues)
    df['Status'] = df['Status'].astype(str)  # Ensure the Status column type is string
    #print(f"DataFrame for project {project_number}:\n{df.head()}")  # DEBUG: Check DataFrame content
    #print(f"DataFrame columns: {df.columns}")  # DEBUG: Verify columns
    project_dataframes[shortened_project_mapping[project_number]] = df

    # Pre-process data into dictionaries for access without indexing issues
    processed_data[shortened_project_mapping[project_number]] = df.to_dict(orient='records')

# Create a timestamped Excel writer
current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
output_filename = f"getProjectsStatusReleaseDefects{current_datetime}.xlsx"

with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    for project_title, df in project_dataframes.items():
        # Sheet names already shortened to 31 characters
        sheet_name = project_title
        df.to_excel(writer, sheet_name=sheet_name, index=False)

# Load the newly created Excel file to modify it
workbook = load_workbook(output_filename)

# Define the formulas for the additional columns
label_status_formula = '''=IFERROR(MID(G2, SEARCH("Status: ", G2) + LEN("Status: "), IF(ISNUMBER(SEARCH(",", G2, SEARCH("Status: ", G2) + LEN("Status: "))), SEARCH(",", G2, SEARCH("Status: ", G2) + LEN("Status: ")) - (SEARCH("Status: ", G2) + LEN("Status: ")), LEN(G2))), "")'''
issuetype_formula = '''=IF(OR(UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "FEATURE", UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "USER STORY", UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "TASK", UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "EPIC", UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "OPERATIONAL", UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)) = "DEFECT"), UPPER(LEFT(G2, FIND(" ", G2 & " ") - 1)), IF(OR(LEFT(G2, 4) = "Pod:", G2 = ""), "", IFERROR(IF(ISERROR(FIND(",", G2)), G2, LEFT(G2, FIND(",", G2) - 1)), G2)))'''
pod_formula = '''=IFERROR(MID(G2, SEARCH("Pod: ", G2) + LEN("Pod: "), SEARCH(",", G2, SEARCH("Pod: ", G2)) - (SEARCH("Pod: ", G2) + LEN("Pod: "))), "")'''
isdefect_formula = '''=IF(ISNUMBER(SEARCH("Defect", G2)), "Defect", "")'''
convertHyperlink = '''=HYPERLINK(B2, TRIM(RIGHT(SUBSTITUTE(B2, "/", REPT(" ", 100)), 100)))'''
font = Font(color="0000FF", underline="single")

# List of shorted sheet names
sheets_to_update = list(shortened_project_mapping.values())

# Ensure each sheet in the workbook is processed
for project_title in workbook.sheetnames:
    if project_title in processed_data:
        sheet = workbook[project_title]
        sheet.sheet_state = 'visible'  # Ensure each sheet is visible

        # Add the headers for new columns
        sheet['H1'] = 'LabelStatus'  # Update header to LabelStatus
        sheet['I1'] = 'LabelIssueType'
        sheet['J1'] = 'Pod'
        sheet['K1'] = 'IsDefect'
        sheet['L1'] = 'Milestone'
        sheet['M1'] = 'GitHub Link '
        sheet['N1'] = 'POD Project '
        sheet['O1'] = 'Status'  # Add Status column

        # Add the formulas to the respective columns for each row of data
        for row, record in enumerate(processed_data[project_title], start=2):
            sheet[f'H{row}'] = label_status_formula.replace('G2', f'G{row}')
            sheet[f'I{row}'] = issuetype_formula.replace('G2', f'G{row}')
            sheet[f'J{row}'] = pod_formula.replace('G2', f'G{row}')
            sheet[f'K{row}'] = isdefect_formula.replace('G2', f'G{row}')
            sheet[f'M{row}'] = convertHyperlink.replace('B2',f'B{row}')
            sheet[f'M{row}'].font=font
            sheet[f'N{row}'] = project_title

            sheet[f'L{row}'] = record.get('Milestone', '')
            status_value = record.get('Status', '')
            #print(f"Writing Status '{status_value}' to row {row}")  # DEBUG: Print status value being written
            sheet[f'O{row}'] = status_value

# Ensure only one active sheet and it's visible
workbook.active = workbook[workbook.sheetnames[0]]
workbook.active.sheet_state = 'visible'

# Save the updated workbook
workbook.save(output_filename)
print(f" The project details extracted to {output_filename} including 'Status'.")

# Collect all issues with milestone values that include "Release 1.6.0", "Release 1.7.0", or "Release 1.8.0"
release_issues = []
defect_issues =[]
for sheet_name in workbook.sheetnames:
    if sheet_name in sheets_to_update:
        sheet = workbook[sheet_name]

        for row in range(2, sheet.max_row + 1):
            milestone = sheet[f'L{row}'].value  # Capture the milestone value
            if milestone in ["Release 1.6.0", "Release 1.7.0", "Release 1.8.0"]:
                issue_data = [sheet[f'{col}{row}'].value for col in 'ABCDEFGHIJKLMNO']
                release_issues.append(issue_data)
               
            label_value = sheet[f'G{row}'].value
            if label_value:  # Ensure the cell_value is not None
                if "Defect" in label_value:
                   defect_issue_data = [sheet[f'{col}{row}'].value for col in 'ABCDEFGHIJKLMNO']
                   defect_issues.append(defect_issue_data)
                   #Defects_sheet[f'K{row}'] = 'Defect' 
                 
# Ensure data was collected
print(f'Collected Release1.8 sheet items total {len(release_issues)} issues for milestones Release 1.6.0, Release 1.7.0, Release 1.8.0')

# Create a new DataFrame for the Release1.8items
columns = ["Title", "URL", "Created At", "Updated At", "State", "Author", "Labels", "LabelStatus", "IssueType", "Pod", "IsDefect","Milestone","GitHub Link","Pod Project","Status"]
release_df = pd.DataFrame(release_issues, columns=columns)

# Add the "Release1.8items" sheet to the workbook
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a') as writer:  # Open existing Excel file to append
    release_df.to_excel(writer, sheet_name="Release1.8items", index=False)

# Create a new DataFrame for the Defects
columns = ["Title", "URL", "Created At", "Updated At", "State", "Author", "Labels", "LabelStatus", "IssueType", "Pod", "IsDefect","Milestone","GitHub Link","Pod Project","Status"]
defect_df = pd.DataFrame(defect_issues, columns=columns)

# Add the "Defects" sheet to the workbook
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a') as writer:  # Open existing Excel file to append
    defect_df.to_excel(writer, sheet_name="Defects", index=False)

# Reload the workbook to ensure changes are captured
workbook = load_workbook(output_filename)

###Code to remove duplicate rows Based on Column B
# Load the "Release1.8items" sheet into a pandas DataFrame
df_release_items = pd.read_excel(output_filename, sheet_name="Release1.8items")
df_defect_items = pd.read_excel(output_filename, sheet_name="Defects")

# Drop duplicates based on the URL column (Column B)
df_release_items.drop_duplicates(subset=["URL"], inplace=True)
df_defect_items.drop_duplicates(subset=["URL"], inplace=True)

# Save the deduplicated DataFrame back to the "Release1.8items" sheet
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_release_items.to_excel(writer, sheet_name="Release1.8items", index=False)
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_defect_items.to_excel(writer, sheet_name="Defects", index=False)

# Reload the workbook to ensure changes are captured
workbook = load_workbook(output_filename)
# Ensure the "Release1.8items" sheet is visible

if "Release1.8items" in workbook.sheetnames:
    release_sheet = workbook["Release1.8items"]
    release_sheet.sheet_state = 'visible'
    # Apply formulas to the Release1.8items sheet
for row in range(2, len(release_df) + 2):
    release_sheet[f'H{row}'] = label_status_formula.replace('G2', f'G{row}')
    release_sheet[f'I{row}'] = issuetype_formula.replace('G2', f'G{row}')
    release_sheet[f'J{row}'] = pod_formula.replace('G2', f'G{row}')
    release_sheet[f'K{row}'] = isdefect_formula.replace('G2', f'G{row}')
    release_sheet[f'M{row}'] = convertHyperlink.replace('B2',f'B{row}')
    release_sheet[f'M{row}'].font = font
# Ensure only one active sheet and it's visible
workbook.active = workbook[workbook.sheetnames[0]]
workbook.active.sheet_state = 'visible'

# Save the updated workbook
workbook.save(output_filename)

# Filter the DataFrame for rows where the value in Column "I" (IssueType) is "Feature"
df_features = df_release_items.copy(sheet_name)

# Save the filtered DataFrame as a new sheet named "Features"
with pd.ExcelWriter(output_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_features.to_excel(writer, sheet_name="Features", index=False)

# Rows to delete (we collect them to avoid changing the structure while iterating)
rows_to_delete = []

# Reload the workbook to ensure changes are captured
workbook = load_workbook(output_filename)

# Ensure the "Features" sheet is visible
if "Features" in workbook.sheetnames:
    features_sheet = workbook["Features"]
    features_sheet.sheet_state = 'visible'
    # Apply formulas to the Features sheet
    for row in range(2, len(df_features) + 2):
        cell_value = features_sheet[f'G{row}'].value  # Get the computed value in cell G{row}
        if cell_value:  # Ensure the cell_value is not None
         if "Feature" in cell_value:
            features_sheet[f'I{row}'] = 'Feature' 
         else:
            rows_to_delete.append(row)
       
        # Delete rows that are not "Feature"
    for row in reversed(rows_to_delete):  # Reverse to avoid index shift issues
      features_sheet.delete_rows(row, 1)
    
    # Second pass: Adjust hyperlinks for remaining rows
    
    for row in range(2, features_sheet.max_row + 1):
            features_sheet[f'M{row}'] = convertHyperlink.replace('B2', f'B{row}')
            features_sheet[f'M{row}'].font = font
        
    # Delete columns J and K from features_sheet
    features_sheet.delete_cols(10)  # Delete column J (10th column)
    features_sheet.delete_cols(10)  # Delete column K (now the 10th column after deletion of previous column J)
    workbook.save(output_filename)

## New code for the Defects
if "Defects" in workbook.sheetnames:
    Defects_sheet = workbook["Defects"]
    Defects_sheet.sheet_state = 'visible'
    # Apply formulas to the Features sheet
    for row in range(2, len(df_defect_items) + 2):
        Defects_sheet[f'M{row}'] = convertHyperlink.replace('B2',f'B{row}')
        Defects_sheet[f'M{row}'].font=font
        cell_value = Defects_sheet[f'G{row}'].value  # Get the computed value in cell G{row}
        if cell_value:  # Ensure the cell_value is not None
         if "Defect" in cell_value:
            Defects_sheet[f'K{row}'] = 'Defect' 
         #else:
         #   rows_to_delete.append(row) 

Defects_sheet.delete_cols(8)  # Delete column H (8th column)
Defects_sheet.delete_cols(8)  # Delete column I (now the 8th column after deletion of previous column J)
Defects_sheet.delete_cols(8)  # Delete column J    

# Ensure only one active sheet and it's visible
workbook.active = workbook[workbook.sheetnames[0]]
workbook.active.sheet_state = 'visible'

# Save the updated workbook
workbook.save(output_filename)

print(f"Issues successfully written to {output_filename} with additional columns including 'Release1.8items' sheet.")
## Begin write to a .MD file

# Load workbook containing the 'Features' sheet
workbook = load_workbook(output_filename, data_only=True)

# Load the 'Features' sheet into a pandas DataFrame
df_features = pd.read_excel(output_filename, sheet_name="Features")

# Prepare to create the Markdown file
md_filename = "Release_Notes.md"

def fetch_issue_body(issue_url, token):
    issue_number = issue_url.split('/')[-1]
    repo_owner = "kpmg-global-technology-and-knowledge"
    repo_name = "Digital-matrix-app"
    api_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/issues/{issue_number}"

    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(api_url, headers=headers)
    if response.status_code == 200:
        issue_data = response.json()
        body = issue_data.get("body", "No description available.")
        #print(f"Original Body: {body}")  # Debugging line to show the original body
        body = issue_data.get("body", "No description available.")
    
    # Remove multiple line breaks
    body = re.sub(r'\n\s*\n', '\n', body).strip()
    
    # Regex pattern to find the "Acceptance Criteria" section more flexibly
    pattern = r'\n\s*##\s*Acceptance Criteria\s*\n'
    split_body = re.split(pattern, body, flags=re.IGNORECASE)
    
    if len(split_body) > 1:
        truncated_body = split_body[0]
    else:
        truncated_body = body  # No "Acceptance Criteria" found, return the full body
    
    # Remove references to 'Charge Code' (case-insensitive)
    charge_code_pattern = r'(?i)charge codes?\b[^\n]*'
    truncated_body = re.sub(charge_code_pattern, '', truncated_body)
    # Remove any residual multiple newlines from charge code removal
    truncated_body = re.sub(r'\n\s*\n', '\n', truncated_body).strip()

    return truncated_body
    
# Create the Markdown content
with open(md_filename, 'w', encoding='utf-8') as md_file:
    md_file.write("# Release Notes\n\n")
    
    for index, row in df_features.iterrows():
        feature_title = row['Title']
        issue_url = row['URL']

         # Fetch the issue body using the URL
        issue_body = fetch_issue_body(issue_url, token)

        # Ensure issue_body is a string
        issue_body = str(issue_body) if issue_body else ""
        # Write the feature title and issue body to the Markdown file with formatting
        md_file.write(f"## **{feature_title}**\n\n")
        md_file.write(f"*{issue_body}*\n\n")
        md_file.write(f"[Issue Link]({issue_url})\n\n")

        defect_body =fetch_defects_content(output_filename)  
        
        # Ensure issue_body is a string
        defect_body = str(defect_body) if defect_body else ""
        # Write the feature title and issue body to the Markdown file with formatting
        md_file.write("## **List of Defects**\n\n")
        md_file.write(f"*{defect_body}*\n\n")
        #md_file.write(f"[Issue Link]({issue_url})\n\n")

print(f"Release notes successfully written to {md_filename}.")